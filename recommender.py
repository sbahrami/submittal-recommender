import pandas as pd
import re
from sklearn.tree import DecisionTreeClassifier
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.pipeline import Pipeline
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# === CONFIGURATION ===
EXCEL_PATH = "C:\Users\SinaB\Metrolinx\Civil Structures - Engineering - Documents\General\TASK TRACKER\Sina_DON'T TOUCH\WORKING_Automated Submittal List.xlsm"     # path to your Excel file
SHEET_NAME = "tbEmailList"             # name of the sheet (or table exported to sheet)
DISCIPLINE_COLS = ["Geotechnical", "Structural", "Tunnel"]

# === 1. LOAD DATA ===
print("Loading data...")
df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)

required_cols = {"Title", "Submittal Revision #"}
if not required_cols.issubset(df.columns):
    raise ValueError("Excel sheet must include 'Title' and 'Submittal Revision #' columns.")

def extract_submittal_code(text: str) -> str:
    """Find S-### or S### pattern anywhere in text (after cleaning)."""
    text = text.upper()
    match = re.search(r"\bS\s?\d+\b", text)
    return match.group(0).replace(" ", "") if match else ""

# Apply text cleaning
df["Title_clean"] = (
    df["Title"]
    .astype(str)
    .str.lower()
    .str.replace(r"[_]", " ", regex=True)
    .str.replace(r"[^a-z0-9\s\-]", "", regex=True)
    .str.replace(r"\s+", " ", regex=True)
    .str.strip()
)

df["Submittal Revision Clean"] = df["Submittal Revision #"].apply(clean_revision_text)
df["SubmittalCode"] = df["Submittal Revision #"].apply(extract_submittal_code)
df["CombinedFeatures"] = df["Title_clean"] + " " + df["SubmittalCode"].fillna("")

# === 3. TRAIN & PREDICT ===
print("Training models and predicting missing labels...")
results = df.copy()
filled_counts = {}

for col in DISCIPLINE_COLS:
    if col not in df.columns:
        print(f"⚠️ Column '{col}' not found in data; skipping.")
        continue

    y = df[col].astype(str).replace("nan", "")
    mask_train = y != ""
    mask_predict = y == ""

    if mask_train.sum() == 0:
        print(f"⚠️ No training data found for '{col}'. Skipping.")
        continue

    # Text vectorization + decision tree pipeline
    model = Pipeline([
        ("vect", CountVectorizer(stop_words="english")),
        ("clf", DecisionTreeClassifier(max_depth=5, random_state=42))
    ])

    model.fit(df.loc[mask_train, "CombinedFeatures"], y[mask_train])
    preds = model.predict(df.loc[mask_predict, "CombinedFeatures"])
    results.loc[mask_predict, col] = preds
    filled_counts[col] = mask_predict.sum()

print("Prediction complete!")
for col, n in filled_counts.items():
    print(f"  → {col}: filled {n} missing values.")

# === 4. WRITE BACK TO EXCEL WITH HIGHLIGHTING ===
print("Writing predictions back to Excel...")

wb = load_workbook(EXCEL_PATH)
ws = wb[SHEET_NAME]

# Find column indices by header name
headers = {cell.value: cell.column for cell in next(ws.iter_rows(min_row=1, max_row=1))}
yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

for col in DISCIPLINE_COLS:
    if col not in headers:
        continue
    col_idx = headers[col]
    for i, row in results.iterrows():
        if df.at[i, col] == "" and results.at[i, col] != "":
            cell = ws.cell(row=i + 2, column=col_idx)  # +2 to skip header
            cell.value = results.at[i, col]
            cell.fill = yellow

wb.save(EXCEL_PATH)
print(f"✅ Results saved and highlighted in '{EXCEL_PATH}'.")

